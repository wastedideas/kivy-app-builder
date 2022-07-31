import os
import platform

if platform.system() == "Windows":
    os.environ["KIVY_GL_BACKEND"] = "angle_sdl2"
import pytz
import typing
from icalendar import Calendar
from kivy.app import App
from kivy.graphics import Rectangle, Color
from kivy.uix import textinput
from kivy.uix.button import Button
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.core.window import Window
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from dateutil.rrule import rruleset, rrulestr
from datetime import datetime, timezone
from pathlib import Path

Window.size = (1200, 1500)

path = "/Users/gmpolunin/Desktop/projects1/icsalendar/testkatyushka12345@gmail.com (2).ics"

MAIN_ICS_FIELDS = [
    "startdt",
    "your_status",
    "all_day",
    "summary",
    "declined_by_organizer",
    "enddt",
]

TZ_EUROPE_MOSCOW = "Europe/Moscow"


class ICSSniffer:
    def __init__(self, file_path: str):
        self.__file_path: str = file_path
        self.__file_endswith: str = ".ics"
        self.__is_file: bool = os.path.isfile(file_path)
        self.__main_path: str

    def get_ics_file_string(self) -> typing.AnyStr:
        try:
            return self.__read_ics_file()
        except Exception as exc:
            raise ICSSnifferException(exc=exc)

    @property
    def is_file(self) -> bool:
        return self.__is_file

    @property
    def main_path(self) -> str:
        return self.__main_path

    def __read_ics_file(self) -> typing.AnyStr:
        if any(
            (
                not os.path.exists(self.__file_path),
                not self.__is_file,
                not self.__file_path.endswith(self.__file_endswith),
            )
        ):
            raise NotADirectoryError

        if self.__is_file:
            self.__main_path = str(Path(self.__file_path).parent.absolute())

        with open(self.__file_path, "r", encoding="utf-8") as ics_file:
            return ics_file.read()


class ICSSnifferException(Exception):
    def __init__(self, exc: typing.Optional[Exception] = None):
        self.exc = exc


class ExcelWorker:
    __workbook: typing.Optional[Workbook] = None

    def __init__(
        self,
        workbook_name: str,
        workbook_extension: str = ".xlsx",
        want_cleared: bool = True,
        date_fields: typing.List[str] = None,
        date_format: str = "DD/MM/YYYY HH:MM:SS",
        sheets_to_create: typing.Tuple = (),
    ):
        self.__workbook_name: str = workbook_name
        self.__workbook_extension: str = workbook_extension
        self.__full_workbook_name: str = self.__workbook_name + self.__workbook_extension
        self.__want_cleared: bool = want_cleared
        self.__date_fields: typing.List[str] = date_fields
        self.__date_format: str = date_format
        self.__sheets_to_create: typing.Tuple = sheets_to_create
        self.__load_or_create_wb(self.__want_cleared, self.__sheets_to_create)

    def fill_workbook(self, all_data: typing.Dict[str, typing.List]):
        try:
            for k, v in all_data.items():
                self.__create_and_fill_ws(sheet_name=k, data_to_fill=v)
            self.__save_and_close_wb()
        except Exception as exc:
            raise ExcelWorkerException(exc=exc)

    @property
    def full_workbook_name(self) -> str:
        return self.__full_workbook_name

    def __load_or_create_wb(self, want_cleared: bool, sheets_to_create: typing.Tuple):
        if want_cleared:
            self.__workbook = Workbook()
            self.__workbook.remove(worksheet=self.__workbook.active)

            if sheets_to_create:
                for sc in sheets_to_create:
                    self.__workbook.create_sheet(title=sc)
        else:
            self.__workbook = load_workbook(self.__full_workbook_name, keep_vba=True)

    def __save_and_close_wb(self):
        self.__workbook.save(self.__full_workbook_name)
        self.__workbook.close()

    def __rename_and_pick_first_ws(self, sheet_name: str) -> Worksheet:
        ws: Worksheet = self.__workbook.worksheets[0]
        ws.title = sheet_name
        return ws

    def __create_named_ws_in_wb(self, sheet_name: str) -> Worksheet:
        return self.__workbook.create_sheet(title=sheet_name)

    def __ws_append_with_data(self, ws: Worksheet, data: typing.List):
        if self.__date_fields:
            for ds in self.__date_fields:
                ws.column_dimensions[ds].number_format = self.__date_format

        for data_row in data:
            ws.append(data_row)

    def __create_and_fill_ws(self, sheet_name: str, data_to_fill: typing.List):
        ws = self.__create_named_ws_in_wb(sheet_name=sheet_name[:30])
        self.__ws_append_with_data(ws=ws, data=data_to_fill)


class ExcelWorkerException(Exception):
    def __init__(self, exc: typing.Optional[Exception] = None):
        self.exc = exc


class ICalendarParser:
    def __init__(self, ics_string: str, mail_to: str, start_date: datetime, end_date: datetime):
        self.__ics_string: str = ics_string
        self.__mail_to: str = mail_to
        self.__start_date: datetime = start_date
        self.__end_date: datetime = end_date
        self.__events: typing.List = []
        self.__cal = filter(
            lambda c: c.name == "VEVENT",
            Calendar.from_ical(self.__ics_string).walk()
        )

    def get_events_from_ics(self):
        for vevent in self.__cal:
            summary = str(vevent.get("summary"))
            raw_star_dt = vevent.get("dtstart").dt
            raw_end_dt = vevent.get("dtend").dt

            organizer = vevent.get("organizer")
            attendee = vevent.get("attendee")
            your_status = str()
            declined_by_organizer = False
            if attendee and organizer:
                for a in attendee:
                    if a.params.get("cn") == organizer.params.get("cn") and a.params.get("partstat") == "DECLINED":
                        declined_by_organizer = True
                    if a.params.get("cn") == self.__mail_to:
                        your_status = a.params.get("partstat")

            end_dt = None
            all_day = False
            if not isinstance(raw_star_dt, datetime):
                all_day = True
                start_dt = self.__date_to_datetime(raw_star_dt)
                if raw_end_dt:
                    end_dt = self.__date_to_datetime(raw_end_dt)
            else:
                start_dt: datetime = raw_star_dt
                end_dt = raw_end_dt

            ex_date = vevent.get("exdate")
            if vevent.get("rrule"):
                reoccur = vevent.get("rrule").to_ical().decode("utf-8")
                for rd in self.__get_recurrent_datetimes(reoccur, start_dt, self.__end_date, ex_date):
                    new_e = {
                        "startdt": rd,
                        "your_status": your_status,
                        "all_day": all_day,
                        "summary": summary,
                        "declined_by_organizer": declined_by_organizer,
                    }
                    if end_dt:
                        new_e["enddt"] = rd + (end_dt - start_dt)
                    self.__append_event(ne=new_e, start=self.__start_date, end=self.__end_date)
            else:
                self.__append_event(
                    {
                        "startdt": start_dt if all_day else start_dt.astimezone(pytz.timezone(TZ_EUROPE_MOSCOW)),
                        "your_status": your_status,
                        "all_day": all_day,
                        "summary": summary,
                        "declined_by_organizer": declined_by_organizer,
                        "enddt": end_dt if all_day else end_dt.astimezone(pytz.timezone(TZ_EUROPE_MOSCOW))
                    },
                    start=self.__start_date,
                    end=self.__end_date,
                )
        self.__events.sort(key=lambda event: event["startdt"])
        self.__setup_none_tzinfo()
        return self.__events

    def __append_event(self, ne, start, end):
        if ne["startdt"] > end:
            return
        if ne["enddt"]:
            if ne["enddt"] < start:
                return

        self.__events.append(ne)

    def __setup_none_tzinfo(self):
        for event in self.__events:
            for ek, ev in event.items():
                if isinstance(ev, datetime):
                    event[ek] = datetime(ev.year, ev.month, ev.day, ev.hour, ev.minute, ev.second, tzinfo=None)

    @staticmethod
    def __get_recurrent_datetimes(recur_rule, start, end, exclusions):
        rules = rruleset()
        first_rule = rrulestr(recur_rule, dtstart=start)
        rules.rrule(first_rule)
        if not isinstance(exclusions, list):
            exclusions = [exclusions]

        for xdt in exclusions:
            try:
                rules.exdate(xdt.dt)
            except AttributeError:
                pass

        dates = []

        for dl in rules.between(start, end):
            dates.append(dl)
        return dates

    @staticmethod
    def __date_to_datetime(dt):
        return datetime(dt.year, dt.month, dt.day, tzinfo=timezone.utc)


class ICalendarParserException(Exception):
    def __init__(self, exc: typing.Optional[Exception] = None):
        self.exc = exc


class ICalendarJob:
    def __init__(self, file_path: str, mail_to: str, start_date: datetime, end_date: datetime):
        self.__file_path: str = file_path
        self.__mail_to: str = mail_to
        self.__start_date: datetime = start_date
        self.__end_date: datetime = end_date
        self.__workbook_name: str = f"{self.__mail_to}-{datetime.isoformat(datetime.now())[:19].replace(':', '-')}"
        self.__ics_sniffer: ICSSniffer = ICSSniffer(
            file_path=self.__file_path,
        )

    def run_sniff_and_write_ics_lines(self):
        ics_sting: typing.AnyStr = self.__ics_sniffer.get_ics_file_string()
        calendar_parser = ICalendarParser(
            ics_string=ics_sting,
            mail_to=self.__mail_to,
            start_date=self.__start_date,
            end_date=self.__end_date,
        )

        ics_list: typing.List = calendar_parser.get_events_from_ics()
        excel_worker = ExcelWorker(
            workbook_name=f"{self.__ics_sniffer.main_path}/{self.__workbook_name}",
            date_fields=["A", "F"],
        )
        if ics_list:
            try:
                data_list = []

                headers_list: typing.List = [h for h in ics_list[0].keys()]
                data_list.append(headers_list)
                for data in ics_list:
                    data_list.append([data[ih] for ih in headers_list])

                excel_worker.fill_workbook(all_data={self.__workbook_name: data_list})
            except ICSSnifferException as e:
                raise ICalendarJobException(msg="Файловая ошибка. Директория/файл не найден.", exc=e.exc)
            except ExcelWorkerException as e:
                raise ICalendarJobException(msg="Внутренняя ошибка работы с Excel.", exc=e.exc)
        return self.__workbook_name


class ICalendarJobException(Exception):
    def __init__(self, exc: typing.Optional[Exception] = None, msg: typing.Optional[str] = None):
        self.exc = exc
        self.msg = msg


class TextInput(textinput.TextInput):
    def __init__(self, **kwargs):
        super(TextInput, self).__init__(**kwargs)
        self.padding_x = [
            self.center[0] - self._get_text_width(max(self._lines, key=len), self.tab_width, self._label_cached) / 2.0,
            0,
        ] if self.text else [self.center[0], 0]
        self.padding_y = [self.height / 2.0 - (self.line_height / 2.0) * len(self._lines), 0]


BLACK = 0, 0, 0, 1
YELLOW = .988, .725, .074, 1


class ICalendarLayout(GridLayout):
    def __init__(self, **kwargs):
        super(ICalendarLayout, self).__init__(**kwargs)

        with self.canvas.before:
            Color(*YELLOW, mode="rgba")
            self.rect = Rectangle(pos=self.pos, size=self.size)
        self.bind(size=self.update_rect)

        self.cols = 1
        self.height = self.minimum_height

        self.general_input = GridLayout(
            cols=2,
            size_hint_y=1,
        )
        self.general_input.add_widget(
            Label(
                text="Путь до директории/файла:",
                color=BLACK,
                bold=True,
                text_size=(None, None),
                font_size="20sp",
            )
        )
        self.directory = TextInput(
            multiline=True,
            hint_text="User/example/path/to/directory",
            is_focusable=True,
        )
        self.general_input.add_widget(self.directory)

        self.general_input.add_widget(
            Label(
                text="Электронная почта:",
                color=BLACK,
                bold=True,
                text_size=(None, None),
                font_size="20sp",
            )
        )
        self.email = TextInput(
            multiline=True,
            hint_text="testkatyushka1234@gmail.com",
            is_focusable=True,
        )
        self.general_input.add_widget(self.email)

        self.general_input.add_widget(
            Label(
                text="Дата начала:",
                color=BLACK,
                bold=True,
                text_size=(None, None),
                font_size="20sp",
            )
        )
        self.start_date = TextInput(
            multiline=True,
            hint_text="2022-07-01",
            is_focusable=True,
        )
        self.general_input.add_widget(self.start_date)

        self.general_input.add_widget(
            Label(
                text="Дата окончания:",
                color=BLACK,
                bold=True,
                text_size=(None, None),
                font_size="20sp",
            )
        )
        self.end_date = TextInput(
            multiline=True,
            hint_text="2022-07-31",
            is_focusable=True,
        )
        self.general_input.add_widget(self.end_date)

        self.add_widget(self.general_input)

        self.submit = Button(
            text="Выполнить",
            background_normal="",
            background_color=BLACK,
            size_hint_y=.3,
            bold=True,
            text_size=(None, None),
            font_size="20sp",
            color=YELLOW,
        )

        self.submit.bind(on_press=self.press)
        self.add_widget(self.submit)

        self.error = Label(
            bold=True,
            text_size=(None, None),
            font_size="20sp",
            padding=[100, 100],
        )
        self.add_widget(self.error)

    def press(self, instance):
        try:
            sd = datetime.strptime(self.start_date.text, "%Y-%m-%d")
            ed = datetime.strptime(self.end_date.text, "%Y-%m-%d")
            icj = ICalendarJob(
                file_path=self.directory.text,
                mail_to=self.email.text,
                start_date=datetime(sd.year, sd.month, sd.day, 0, 0, 1, tzinfo=timezone.utc),
                end_date=datetime(ed.year, ed.month, ed.day, 0, 0, 1, tzinfo=timezone.utc),
            )
            new_excel = icj.run_sniff_and_write_ics_lines()
            self.error.color = "green"
            self.error.text = f"Успешно:\n{new_excel}"
        except ICalendarJobException as e:
            self.error.color = "red"
            self.error.text = f"{e.msg}\n{e.exc}"
        except Exception as e:
            self.error.color = "red"
            self.error.text = f"{e.__traceback__}"

    def update_rect(self, *args):
        self.rect.pos = self.pos
        self.rect.size = self.size


class ICalendarApp(App):
    icon = "calendar_ico.png"
    title = "Calendar 🗓"

    def build(self):
        return ICalendarLayout()


if __name__ == "__main__":
    ICalendarApp().run()
